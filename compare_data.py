import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def compare_and_merge(path1, path2, output_path):
    # Get all Excel files in the given paths
    files1 = [f for f in os.listdir(path1) if f.endswith('.xlsx')]
    files2 = [f for f in os.listdir(path2) if f.endswith('.xlsx')]

    # Iterate through common files
    for file in set(files1).intersection(files2):
        file1_path = os.path.join(path1, file)
        file2_path = os.path.join(path2, file)

        # Load Excel files into pandas dataframes
        df1 = pd.read_excel(file1_path, engine='openpyxl', sheet_name=None, header=None)
        df2 = pd.read_excel(file2_path, engine='openpyxl', sheet_name=None, header=None)

        # Create a new workbook
        merged_workbook = Workbook()

        # Iterate through sheets
        for sheet_name in set(df1.keys()).union(df2.keys()):
            # Get dataframes for each sheet
            sheet_df1 = df1.get(sheet_name, pd.DataFrame())
            sheet_df2 = df2.get(sheet_name, pd.DataFrame())

            # Skip if both dataframes are empty
            if sheet_df1.empty and sheet_df2.empty:
                continue

            # Create a new sheet in the merged workbook
            merged_sheet = merged_workbook.create_sheet(title=sheet_name)

            # Write data from the first dataset with header
            if not sheet_df1.empty:
                merged_sheet.append([f'\n{sheet_name} from File 1'])
                for row in sheet_df1.iterrows():
                    merged_sheet.append(row[1].tolist())

            # Write an empty row as a separator
            merged_sheet.append([])

            # Write data from the second dataset with header
            if not sheet_df2.empty:
                merged_sheet.append([f'\n{sheet_name} from File 2'])
                for row in sheet_df2.iterrows():
                    merged_sheet.append(row[1].tolist())

            # Identify unique values and highlight in the merged sheet
            highlight_unique_values(merged_sheet, sheet_df1, sheet_df2)

        # Save the merged workbook
        merged_workbook.save(os.path.join(output_path, f'Merged_{file}'))

def highlight_unique_values(merged_sheet, sheet_df1, sheet_df2):
    for row_idx, row in enumerate(merged_sheet.iter_rows(min_row=2, max_row=merged_sheet.max_row, min_col=1, max_col=merged_sheet.max_column), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if (
                sheet_df1.iat[row_idx - 2, col_idx - 1] != sheet_df2.iat[row_idx - 2, col_idx - 1]
            ):
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

if __name__ == "__main__":
    # Replace these paths with your actual paths
    path1 = 'path/to/excels1'
    path2 = 'path/to/excels2'
    output_path = 'path/to/output'

    compare_and_merge(path1, path2, output_path)
